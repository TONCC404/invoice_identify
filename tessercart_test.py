import json
import os
from PIL import Image
import pytesseract
from pdf2image import convert_from_path
import easyocr
import base64
import argparse

pytesseract.pytesseract.tesseract_cmd = r'D:\tessercart\tesseract.exe'

def tessercart_extract_text_from_image(image_path):
    image = Image.open(image_path)
    text = pytesseract.image_to_string(image, lang='eng')
    return text

def easy_ocr(image_path):
    """result with higher precision than tessercart"""
    reader = easyocr.Reader(['en', 'ch_tra'])
    results = reader.readtext(image_path)
    for (bbox, text, prob) in results:
        print(f"Detected: {text} (Confidence: {prob:.2f})")

def extract_text_from_pdf(pdf_path):
    """识别PDF文件的文字"""
    images = convert_from_path(pdf_path)
    text = ""
    for img in images:
      text += pytesseract.image_to_string(img, lang='eng') + "\n"
    return text


def encode_image_to_base64(file_path):
  with open(file_path, "rb") as image_file:
    base64_bytes = base64.b64encode(image_file.read())
    base64_string = base64_bytes.decode('utf-8')
  return f"data:image/png;base64,{base64_string}"

from typing import List
from pydantic import BaseModel
class Item(BaseModel):
  product_id: str
  matched_name: str
  original_input: str
  quantity:int
  match_score: float

class Output(BaseModel):
  customer_name: str
  order_date: str
  items: List[Item]
  status: int

def clean_and_parse(res):
  if res.startswith('```json'):
    res = res[len('```json'):].strip()
    if res.endswith('```'):
      res = res[:-3].strip()
  return json.loads(res)


def llm(path, token = None):
    import requests
    try:
        example_json= {
          "customer_name": "Tony Wang",
          "order_date": "2025-03-25",
          "items": [
            {
              "product_id": "P001",
              "matched_name": "Apple MacBook Air M2",
              "original_input": "mac air laptop",
              "quantity": 2,
              "match_score": 0.92
            },
            {
              "product_id": "P003",
              "matched_name": "Dell 24'' Monitor",
              "original_input": "24 dell screen",
              "quantity": 1,
              "match_score": 0.87
            }
          ],
          "status": "completed"
        }
        url = "https://api.ap.siliconflow.com/v1/chat/completions"
        image_url = encode_image_to_base64(path)
        payload = {
          "model": "Qwen/Qwen2.5-VL-32B-Instruct",
          "messages": [
            {
              "role": "system",
              "content": [{"type": "text", "text": f"请根据图片内容输出json结构得格式，格式例子如下：{json.dumps(example_json)}"}]
            },
            {
              "role": "user",
              "content": [
                {"type": "text", "text": "请根据图片内容输出json结构得格式"},
                {"type":"image_url",
                 "image_url": {"url":image_url,"detail":"low"}
                }
              ]
            }
          ],
          "stream": False,
          "max_tokens": 8192,
          "stop": None,
          "temperature": 0.4,
          "top_p": 0.7,
          "top_k": 50,
          "frequency_penalty": 0.5,
          "n": 1,
          "response_format": {"type": "text"}
        }
        headers = {
          "Authorization": f"Bearer {token}",
          "Content-Type": "application/json"
        }

        response = requests.request("POST", url, json=payload, headers=headers)
        res = json.loads(response.text)
        res = res['choices'][0]['message']['content']
        return clean_and_parse(res)
    except Exception as e:
        print(f"error is:{e}")


def recognize_text(file_path, use_choice=None, token = None):
    """根据文件类型选择识别方式"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
        if use_choice == 'tessercart':
            return tessercart_extract_text_from_image(image_path=file_path)
        elif use_choice == 'easyOCR':
            return easy_ocr(image_path=file_path)
        else:
            return llm(path = file_path, token = token)
    elif ext == '.pdf':
        return extract_text_from_pdf(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

def parse_xlsx():
  from openpyxl import load_workbook
  wb = load_workbook('order.xlsx')
  ws = wb.active
  name = []
  for row in ws.iter_rows(values_only=True):
      name.append(row[1])
  return name[1:]

def main():
    # todo add rag
    # correct_name_list = parse_xlsx()
    parser = argparse.ArgumentParser(description="This is a terminal app")
    parser.add_argument('--path', type=str, help='path for png,jpg,pdf file')
    parser.add_argument('--method', type=str, choices=['tessercart', 'easyOCR', 'llm'], help='method for parse pic')
    parser.add_argument('--token', type=str, help='token for llm')


    args = parser.parse_args()
    if args.path and args.method:
        result = recognize_text(file_path = args.path, use_choice =args.method, token = args.token)
        print(result)

if __name__ == "__main__":
    main()
